#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Веб-приложение для инвентаризации складских мест.
"""

import asyncio
import logging
import os
import sys

from dotenv import load_dotenv
load_dotenv()
from datetime import datetime, timedelta
from pathlib import Path

from flask import Flask, g, jsonify, render_template, request, send_file, redirect
from flask_cors import CORS
from werkzeug.exceptions import BadRequest
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

import psycopg2
from psycopg2.extras import RealDictCursor

from scraper.database import DatabaseConfig

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)

# Инициализация Flask
app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False

# Настройка CORS для Safari и других браузеров
CORS(app, resources={
    r"/*": {
        "origins": "*",
        "methods": ["GET", "POST", "OPTIONS"],
        "allow_headers": ["Content-Type"],
        "supports_credentials": True
    }
})

# Конфигурация БД (из переменных окружения в Docker/продакшене, иначе значения по умолчанию)
DB_CONFIG = DatabaseConfig(
    host=os.environ.get("DB_HOST", "31.207.77.167"),
    port=int(os.environ.get("DB_PORT", "5432")),
    database=os.environ.get("DB_NAME", "botdb"),
    user=os.environ.get("DB_USER", "aperepechkin"),
    password=os.environ.get("DB_PASSWORD", "password"),
)

# Глобальное подключение к БД (синхронное). На Vercel не используется — соединение на каждый запрос.
db_connection = None
IS_VERCEL = os.environ.get("VERCEL") == "1"


@app.route('/sw.js')
def service_worker():
    """Service Worker для PWA (scope / через Service-Worker-Allowed)."""
    from flask import send_from_directory
    resp = send_from_directory(app.static_folder, 'sw.js', mimetype='application/javascript')
    resp.headers['Service-Worker-Allowed'] = '/'
    return resp


@app.route('/api/health', methods=['GET'])
def health_check():
    """Простой health-check для фронта."""
    try:
        # Лёгкая проверка БД: один SELECT 1
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute("SELECT 1")
            cur.fetchone()
        return jsonify({'ok': True}), 200
    except Exception as exc:
        logger.warning("Health-check failed: %s", exc)
        return jsonify({'ok': False, 'error': 'Проблема соединения с сервером'}), 503


def get_db():
    """Получить подключение к БД. На Vercel — одно соединение на запрос (хранится в g), после ответа закрывается."""
    if IS_VERCEL:
        if getattr(g, "db_conn", None) is None or g.db_conn.closed:
            g.db_conn = psycopg2.connect(
                host=DB_CONFIG.host,
                port=DB_CONFIG.port,
                database=DB_CONFIG.database,
                user=DB_CONFIG.user,
                password=DB_CONFIG.password,
                cursor_factory=RealDictCursor,
            )
            ensure_tasks_table()
        return g.db_conn
    global db_connection
    if db_connection is None or db_connection.closed:
        db_connection = psycopg2.connect(
            host=DB_CONFIG.host,
            port=DB_CONFIG.port,
            database=DB_CONFIG.database,
            user=DB_CONFIG.user,
            password=DB_CONFIG.password,
            cursor_factory=RealDictCursor
        )
        logger.info("Подключение к БД установлено")
        ensure_tasks_table()
    return db_connection


@app.teardown_appcontext
def close_db_on_vercel(exception=None):
    """На Vercel закрываем соединение с БД после каждого запроса."""
    if IS_VERCEL and getattr(g, "db_conn", None) is not None:
        try:
            if not g.db_conn.closed:
                g.db_conn.close()
        except Exception as e:
            logger.warning("Ошибка при закрытии соединения: %s", e)
        g.db_conn = None


def safe_rollback(conn):
    """Безопасный rollback транзакции."""
    if conn and not conn.closed:
        try:
            conn.rollback()
        except Exception as e:
            logger.error("Ошибка при rollback: %s", e)


def ensure_tasks_table():
    """Создаёт таблицы active_tasks, inventory_results и связанные структуры."""
    conn = None
    try:
        conn = get_db()
        with conn.cursor() as cur:
            # Таблица активных заданий
            cur.execute("""
                CREATE TABLE IF NOT EXISTS active_tasks (
                    task_id SERIAL PRIMARY KEY,
                    zone_prefix VARCHAR(50) NOT NULL,
                    badge VARCHAR(100) NOT NULL,
                    assigned_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    expires_at TIMESTAMP NOT NULL,
                    status VARCHAR(20) DEFAULT 'active'
                )
            """)
            
            # Таблица результатов инвентаризации
            cur.execute("""
                CREATE TABLE IF NOT EXISTS inventory_results (
                    result_id SERIAL PRIMARY KEY,
                    badge VARCHAR(100) NOT NULL,
                    place_cod BIGINT NOT NULL,
                    place_name VARCHAR(255),
                    plt_id VARCHAR(50),
                    qty_shk_db INTEGER,
                    qty_shk_fact INTEGER,
                    status VARCHAR(50),
                    has_discrepancy BOOLEAN DEFAULT FALSE,
                    photo_data BYTEA,
                    photo_filename VARCHAR(255),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Таблица дополнительных фото к результатам инвентаризации
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS inventory_result_photos (
                    photo_id INTEGER PRIMARY KEY,
                    result_id INTEGER NOT NULL REFERENCES inventory_results(result_id) ON DELETE CASCADE,
                    badge VARCHAR(100) NOT NULL,
                    place_cod BIGINT,
                    photo_data BYTEA NOT NULL,
                    photo_filename VARCHAR(255) NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
            # Гарантируем наличие sequence и auto-increment для photo_id
            cur.execute(
                "CREATE SEQUENCE IF NOT EXISTS inventory_result_photos_photo_id_seq"
            )
            cur.execute(
                """
                ALTER TABLE inventory_result_photos
                ALTER COLUMN photo_id SET DEFAULT nextval('inventory_result_photos_photo_id_seq')
                """
            )
            cur.execute(
                """
                ALTER SEQUENCE inventory_result_photos_photo_id_seq
                OWNED BY inventory_result_photos.photo_id
                """
            )
            
            # Таблица сессий пользователей
            cur.execute("""
                CREATE TABLE IF NOT EXISTS user_sessions (
                    session_id SERIAL PRIMARY KEY,
                    badge VARCHAR(100) NOT NULL,
                    login_time TIMESTAMP NOT NULL,
                    logout_time TIMESTAMP,
                    total_scanned INTEGER DEFAULT 0,
                    with_discrepancy INTEGER DEFAULT 0,
                    no_discrepancy INTEGER DEFAULT 0,
                    session_duration INTEGER,
                    is_active BOOLEAN DEFAULT TRUE
                )
            """)
            
            # Таблица отчетов (для админ-панели)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS reports (
                    report_id SERIAL PRIMARY KEY,
                    badge VARCHAR(100) NOT NULL,
                    file_data BYTEA NOT NULL,
                    filename VARCHAR(255) NOT NULL,
                    total_scanned INTEGER DEFAULT 0,
                    with_discrepancy INTEGER DEFAULT 0,
                    no_discrepancy INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    downloaded_at TIMESTAMP,
                    downloaded_by VARCHAR(100)
                )
            """)

            # Таблица ревизий качества
            cur.execute("""
                CREATE TABLE IF NOT EXISTS quality_reviews (
                    review_id SERIAL PRIMARY KEY,
                    zone_prefix VARCHAR(50) NOT NULL,
                    reviewer VARCHAR(100),
                    status VARCHAR(30) DEFAULT 'planned',
                    summary TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Таблица тикетов/инцидентов
            cur.execute("""
                CREATE TABLE IF NOT EXISTS tickets (
                    ticket_id SERIAL PRIMARY KEY,
                    badge VARCHAR(100) NOT NULL,
                    place_cod BIGINT,
                    description TEXT NOT NULL,
                    priority VARCHAR(20) DEFAULT 'medium',
                    status VARCHAR(20) DEFAULT 'open',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    resolved_at TIMESTAMP,
                    resolver VARCHAR(100)
                )
            """)
            
            # Индексы
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_active_tasks_zone 
                ON active_tasks(zone_prefix, status)
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_active_tasks_expires 
                ON active_tasks(expires_at)
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_inventory_results_badge 
                ON inventory_results(badge)
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_inventory_results_date 
                ON inventory_results(created_at)
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_inventory_results_discrepancy 
                ON inventory_results(has_discrepancy)
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_inventory_result_photos_result 
                ON inventory_result_photos(result_id)
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_inventory_result_photos_place 
                ON inventory_result_photos(place_cod)
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_user_sessions_badge 
                ON user_sessions(badge)
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_user_sessions_active 
                ON user_sessions(is_active)
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_reports_badge 
                ON reports(badge)
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_reports_created 
                ON reports(created_at DESC)
            """)
            
            conn.commit()
            logger.info("✅ Таблицы БД созданы")
    except Exception as e:
        safe_rollback(conn)
        logger.error("Ошибка создания таблиц: %s", e)


def save_report(badge, file_data, filename, total_scanned, with_discrepancy, no_discrepancy):
    """Сохранить отчет в БД для админ-панели."""
    conn = None
    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO reports 
                (badge, file_data, filename, total_scanned, with_discrepancy, no_discrepancy)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING report_id
            """, (badge, psycopg2.Binary(file_data), filename, total_scanned, with_discrepancy, no_discrepancy))
            
            report_id = cur.fetchone()['report_id']
            conn.commit()
            
        logger.info(f"Отчет сохранен: {filename} (ID: {report_id})")
        return report_id
    
    except Exception as e:
        safe_rollback(conn)
        logger.error(f"Ошибка сохранения отчета: {e}")
        return None


def cleanup_expired_tasks():
    """Очищает просроченные задания."""
    conn = None
    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE active_tasks 
                SET status = 'expired' 
                WHERE expires_at < CURRENT_TIMESTAMP 
                    AND status = 'active'
            """)
            expired_count = cur.rowcount
            conn.commit()
            
            if expired_count > 0:
                logger.info("Очищено просроченных заданий: %d", expired_count)
    except Exception as e:
        safe_rollback(conn)
        logger.error("Ошибка очистки просроченных заданий: %s", e)


def reserve_zone(zone_prefix: str, badge: str, hours: int = 2) -> bool:
    """
    Резервирует зону для сотрудника.
    
    Returns:
        True если зона зарезервирована, False если занята
    """
    conn = None
    try:
        conn = get_db()
        with conn.cursor() as cur:
            # Проверяем, не занята ли уже эта зона
            cur.execute("""
                SELECT badge 
                FROM active_tasks 
                WHERE zone_prefix = %s 
                    AND status = 'active'
                    AND expires_at > CURRENT_TIMESTAMP
            """, (zone_prefix,))
            
            existing = cur.fetchone()
            
            if existing:
                logger.warning("Зона %s уже занята сотрудником %s", zone_prefix, existing['badge'])
                return False
            
            # Резервируем зону
            cur.execute("""
                INSERT INTO active_tasks (zone_prefix, badge, expires_at)
                VALUES (%s, %s, CURRENT_TIMESTAMP + INTERVAL '%s hours')
            """, (zone_prefix, badge, hours))
            conn.commit()
            
            logger.info("Зона %s зарезервирована для %s на %d часов", zone_prefix, badge, hours)
            return True
            
    except Exception as e:
        safe_rollback(conn)
        logger.error("Ошибка резервирования зоны: %s", e)
        return False


def get_occupied_zones() -> set:
    """Получает список занятых зон."""
    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT DISTINCT zone_prefix 
                FROM active_tasks 
                WHERE status = 'active'
                    AND expires_at > CURRENT_TIMESTAMP
            """)
            
            zones = {row['zone_prefix'] for row in cur.fetchall()}
            logger.info("Занятых зон: %d", len(zones))
            return zones
            
    except Exception as e:
        logger.error("Ошибка получения занятых зон: %s", e)
        return set()


@app.route('/')
def index():
    """Главная страница - форма авторизации."""
    return render_template('login.html')


@app.route('/work')
def work():
    """Рабочая форма инвентаризации."""
    employee_badge = request.cookies.get('employee_badge')
    if not employee_badge:
        return redirect('/')
    return render_template('work.html', employee_badge=employee_badge)


@app.route('/admin')
def admin():
    """Страница входа в админ-панель. Если уже авторизован — редирект в дашборд."""
    admin_badge = request.cookies.get('admin_badge')
    if admin_badge and admin_badge == 'ADMIN':
        return redirect('/admin/dashboard')
    return render_template('admin_login.html')


@app.route('/admin_login')
def admin_login_legacy():
    """Старый путь для обратной совместимости."""
    return redirect('/admin')


@app.route('/admin/dashboard')
def admin_dashboard():
    """Дашборд админ-панели (защищенный)."""
    # Проверка авторизации
    admin_badge = request.cookies.get('admin_badge')
    if not admin_badge or admin_badge != 'ADMIN':
        return redirect('/admin')
    return render_template('admin.html')


@app.route('/api/auth', methods=['POST'])
def auth():
    """Авторизация по бэйджу сотрудника или администратора."""
    data = request.get_json()
    badge = data.get('badge', '').strip()
    password = data.get('password', '').strip()
    
    if not badge:
        return jsonify({'error': 'Бэйдж не указан'}), 400
    
    # Проверка на администратора
    ADMIN_BADGE = 'ADMIN'
    ADMIN_PASSWORD = 'admin123'
    
    if badge == ADMIN_BADGE:
        # Это попытка входа администратора
        if not password:
            # Нужен пароль
            return jsonify({
                'success': False,
                'require_password': True,
                'message': 'Введите пароль администратора'
            }), 200
        
        # Проверка пароля
        if password != ADMIN_PASSWORD:
            return jsonify({'error': 'Неверный пароль администратора'}), 401
        
        # Успешный вход администратора
        response = jsonify({
            'success': True,
            'is_admin': True,
            'badge': badge,
            'redirect': '/admin/dashboard',
            'timestamp': datetime.now().isoformat()
        })
        response.set_cookie('admin_badge', badge, max_age=28800, httponly=True)
        logger.info("Вход администратора: badge=%s", badge)
        return response
    
    # Обычный сотрудник
    if len(badge) < 3:
        return jsonify({'error': 'Некорректный бэйдж'}), 400
    
    logger.info("Авторизация сотрудника: badge=%s", badge)
    
    response = jsonify({
        'success': True,
        'is_admin': False,
        'badge': badge,
        'redirect': '/work',
        'timestamp': datetime.now().isoformat()
    })
    response.set_cookie('employee_badge', badge, max_age=28800, httponly=True)
    response.set_cookie('admin_badge', '', max_age=0)
    logger.info("Вход сотрудника: badge=%s", badge)
    return response


@app.route('/api/user/stats/<badge>', methods=['GET'])
def get_user_stats(badge):
    """Получить статистику пользователя. Если передан since (timestamp в мс) — только данные смены с этого момента."""
    try:
        since_ts = request.args.get('since', type=lambda x: int(x) if x and str(x).isdigit() else None)
        since_sql = ""
        since_arg = ()
        if since_ts and since_ts > 0:
            since_sql = " AND created_at >= to_timestamp(%s::double precision / 1000)"
            since_arg = (since_ts,)

        conn = get_db()
        with conn.cursor() as cur:
            # Общая статистика за все время (без фильтра по смене)
            cur.execute("""
                SELECT 
                    COUNT(*) as total_scanned,
                    SUM(CASE WHEN has_discrepancy THEN 1 ELSE 0 END) as with_discrepancy,
                    SUM(CASE WHEN NOT has_discrepancy THEN 1 ELSE 0 END) as no_discrepancy
                FROM inventory_results
                WHERE badge = %s
            """, (badge,))
            overall = dict(cur.fetchone() or {})

            # Статистика по сессиям
            cur.execute("""
                SELECT 
                    session_id,
                    login_time,
                    logout_time,
                    total_scanned,
                    with_discrepancy,
                    no_discrepancy,
                    session_duration,
                    is_active
                FROM user_sessions
                WHERE badge = %s
                ORDER BY login_time DESC
                LIMIT 10
            """, (badge,))
            sessions = []
            for row in cur.fetchall():
                sessions.append({
                    'session_id': row['session_id'],
                    'login_time': row['login_time'].isoformat() if row['login_time'] else None,
                    'logout_time': row['logout_time'].isoformat() if row['logout_time'] else None,
                    'total_scanned': row['total_scanned'],
                    'with_discrepancy': row['with_discrepancy'],
                    'no_discrepancy': row['no_discrepancy'],
                    'session_duration': row['session_duration'],
                    'is_active': row['is_active']
                })

            # Статистика за сегодня (или за смену, если передан since)
            cur.execute("""
                SELECT 
                    COUNT(*) as today_scanned,
                    SUM(CASE WHEN has_discrepancy THEN 1 ELSE 0 END) as today_discrepancy
                FROM inventory_results
                WHERE badge = %s 
                AND created_at >= CURRENT_DATE
                """ + since_sql,
                (badge,) + since_arg
            )
            today = dict(cur.fetchone() or {})

            # Последние 5 сканов (за смену, если передан since)
            cur.execute("""
                SELECT place_cod, place_name, status, has_discrepancy, photo_filename, created_at
                FROM inventory_results
                WHERE badge = %s
                """ + (since_sql if since_sql else "") + """
                ORDER BY created_at DESC
                LIMIT 5
            """, (badge,) + since_arg if since_sql else (badge,))
            last_places = []
            for row in cur.fetchall():
                last_places.append({
                    'place_cod': row['place_cod'],
                    'place_name': row['place_name'],
                    'status': row['status'],
                    'has_discrepancy': row['has_discrepancy'],
                    'has_photo': bool(row.get('photo_filename')),
                    'created_at': row['created_at'].isoformat() if row['created_at'] else None,
                    'updated_at': row['created_at'].isoformat() if row['created_at'] else None,
                })

        return jsonify({
            'success': True,
            'badge': badge,
            'overall': overall,
            'today': today,
            'sessions': sessions,
            'last_places': last_places
        })
    
    except Exception as e:
        logger.exception("Ошибка получения статистики пользователя")
        return jsonify({'error': str(e)}), 500


@app.route('/api/user/daily-stats/<badge>', methods=['GET'])
def get_user_daily_stats(badge):
    """Динамика сканов пользователя за последние 7 дней."""
    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    DATE(created_at) as date,
                    COUNT(*) as total,
                    SUM(CASE WHEN has_discrepancy THEN 1 ELSE 0 END) as errors,
                    SUM(CASE WHEN NOT has_discrepancy THEN 1 ELSE 0 END) as ok
                FROM inventory_results
                WHERE badge = %s AND created_at >= CURRENT_DATE - INTERVAL '7 days'
                GROUP BY DATE(created_at)
                ORDER BY date
            """, (badge,))
            rows = cur.fetchall()
        return jsonify({
            'success': True,
            'daily': [{
                'date': r['date'].isoformat(),
                'total': r['total'],
                'errors': r['errors'],
                'ok': r['ok']
            } for r in rows]
        })
    except Exception as e:
        logger.exception("Ошибка daily stats")
        return jsonify({'error': str(e)}), 500


@app.route('/api/user/history', methods=['GET'])
def get_user_history():
    """Получить историю сканов сотрудника с фильтром по датам."""
    badge = request.args.get('badge', '').strip()
    date_from = request.args.get('from')
    date_to = request.args.get('to')

    if not badge:
        return jsonify({'error': 'Не указан badge'}), 400

    # Если даты не заданы — берём сегодня
    try:
        conn = get_db()
        with conn.cursor() as cur:
            query = """
                SELECT 
                    result_id,
                    place_cod,
                    place_name,
                    qty_shk_db,
                    qty_shk_fact,
                    status,
                    has_discrepancy,
                    photo_filename,
                    created_at
                FROM inventory_results
                WHERE badge = %s
            """
            params = [badge]

            if date_from:
                query += " AND created_at >= %s"
                params.append(date_from)
            if date_to:
                query += " AND created_at < %s::date + INTERVAL '1 day'"
                params.append(date_to)

            query += " ORDER BY created_at DESC LIMIT 500"

            cur.execute(query, params)
            rows = cur.fetchall()

        history = []
        for row in rows:
            history.append(
                {
                    "id": row["result_id"],
                    "place_cod": row["place_cod"],
                    "place_name": row["place_name"],
                    "qty_db": row["qty_shk_db"],
                    "qty_fact": row["qty_shk_fact"],
                    "status": row["status"],
                    "has_discrepancy": row["has_discrepancy"],
                    "has_photo": bool(row["photo_filename"]),
                    "created_at": row["created_at"].isoformat() if row["created_at"] else None,
                }
            )

        return jsonify({"success": True, "badge": badge, "history": history})
    except Exception as e:
        logger.exception("Ошибка получения истории сканов пользователя")
        return jsonify({"error": str(e)}), 500


@app.route('/api/user/history/export', methods=['GET'])
def export_user_history():
    """Скачать все сканы сотрудника со статусом не OK (или с расхождением) за период."""
    badge = request.args.get('badge', '').strip()
    date_from = request.args.get('from')
    date_to = request.args.get('to')

    if not badge:
        return jsonify({'error': 'Не указан badge'}), 400

    try:
        conn = get_db()
        with conn.cursor() as cur:
            query = """
                SELECT 
                    ir.result_id,
                    ir.created_at,
                    ir.place_cod,
                    ir.place_name,
                    ir.qty_shk_db,
                    ir.qty_shk_fact,
                    ir.status,
                    ir.has_discrepancy,
                    ir.photo_filename,
                    ir.discrepancy_reason,
                    ir.comment,
                    wp.storage_type,
                    wp.box_type,
                    wp.category,
                    wp.dimensions,
                    COALESCE(wp.floor, wp2.floor) AS floor,
                    COALESCE(wp.row_num, wp2.row_num) AS row_num,
                    COALESCE(wp.section, wp2.section) AS section
                FROM inventory_results ir
                LEFT JOIN warehouse_places wp ON wp.mx_id = ir.place_cod
                LEFT JOIN warehouse_places wp2 ON UPPER(TRIM(wp2.mx_code)) = UPPER(TRIM(ir.place_name))
                    AND ir.place_name IS NOT NULL AND ir.place_name != ''
                WHERE ir.badge = %s
                  AND (LOWER(COALESCE(ir.status, '')) <> 'ok' OR ir.has_discrepancy = TRUE)
            """
            params = [badge]

            if date_from:
                query += " AND ir.created_at >= %s"
                params.append(date_from)
            if date_to:
                query += " AND ir.created_at < %s::date + INTERVAL '1 day'"
                params.append(date_to)

            query += " ORDER BY ir.created_at DESC"

            cur.execute(query, params)
            rows = cur.fetchall()

        if not rows:
            return jsonify({'error': 'Нет записей со статусом не OK за выбранный период'}), 404

        def _format_mx_type(storage_type, box_type=None, category=None, dimensions=None):
            """Только Полка или Короб по данным из отчёта Вместимость и заполненность."""
            for val in (storage_type, box_type, category):
                if not val:
                    continue
                s = str(val).lower()
                if "короб" in s or "box" in s:
                    return "Короб"
                if "полка" in s or "shelf" in s or "стеллаж" in s:
                    return "Полка"
            if dimensions:
                try:
                    parts = str(dimensions).replace("х", "x").replace("Х", "x").split("x")
                    nums = [int(p.strip()) for p in parts if p.strip().isdigit()]
                    if nums:
                        if max(nums) > 900:
                            return "Полка"
                        return "Короб"
                except (ValueError, TypeError):
                    pass
            if storage_type or box_type or category:
                return "Короб"
            return ""

        def _status_label(status):
            if not status:
                return ""
            s = (status or "").lower()
            if s == "ok":
                return "Совпадает"
            if s == "error":
                return "Ошибка"
            if s == "shelf_error":
                return "Поломалось"
            if s == "recount":
                return "Пересорт"
            if s == "missing":
                return "Отсутствует"
            return status

        def _error_description(reason, comment):
            parts = []
            if reason:
                parts.append(str(reason).strip())
            if comment:
                parts.append(f"Коммент.: {str(comment).strip()}")
            return " | ".join(parts) if parts else "—"

        def _parse_mx_code(mx_code):
            """Извлечь этаж, ряд, секцию из кода МХ формата 36.02.40.140.06.03."""
            if not mx_code or not isinstance(mx_code, str):
                return None, None, None
            parts = str(mx_code).strip().split(".")
            try:
                floor = int(parts[1]) if len(parts) >= 2 else None
                row_num = int(parts[2]) if len(parts) >= 3 else None
                section = int(parts[3]) if len(parts) >= 4 else None
                return floor, row_num, section
            except (ValueError, IndexError):
                return None, None, None

        # Формируем Excel-файл
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.drawing.image import Image as XlImage
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Проблемные сканы"

        headers = [
            "Дата/время", "Этаж", "Ряд", "Секция",
            "Код МХ", "ID места",
            "Статус", "Что за ошибка (причина)", "Фото"
        ]
        ws.append(headers)

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
            cell.fill = header_fill

        photo_col = len(headers)
        row_height = 75
        img_max_height = 60
        MAX_EMBEDDED_PHOTOS = 100

        for excel_row_idx, row in enumerate(rows, start=2):
            created = row["created_at"]
            has_photo = bool(row.get("photo_filename"))
            photo_data = None
            if has_photo and row.get("result_id") and (excel_row_idx - 2) < MAX_EMBEDDED_PHOTOS:
                with conn.cursor() as cur_ph:
                    cur_ph.execute(
                        "SELECT photo_data FROM inventory_results WHERE result_id = %s AND photo_data IS NOT NULL LIMIT 1",
                        (row["result_id"],),
                    )
                    ph_row = cur_ph.fetchone()
                    if ph_row:
                        photo_data = ph_row.get("photo_data")
                    if not photo_data:
                        cur_ph.execute(
                            "SELECT photo_data FROM inventory_result_photos WHERE result_id = %s AND photo_data IS NOT NULL LIMIT 1",
                            (row["result_id"],),
                        )
                        ph_row = cur_ph.fetchone()
                        if ph_row:
                            photo_data = ph_row.get("photo_data")

            photo_cell_value = "есть" if has_photo else ""
            if has_photo and photo_data and (excel_row_idx - 2) < MAX_EMBEDDED_PHOTOS:
                try:
                    raw = photo_data
                    if hasattr(raw, "tobytes"):
                        raw = raw.tobytes()
                    elif not isinstance(raw, bytes):
                        raw = bytes(raw)
                    if raw:
                        img_io = BytesIO(raw)
                        img_io.seek(0)
                        xl_img = XlImage(img_io)
                        if xl_img.height and xl_img.height > img_max_height:
                            ratio = img_max_height / xl_img.height
                            xl_img.height = img_max_height
                            xl_img.width = int(xl_img.width * ratio)
                        xl_img.anchor = f"{get_column_letter(photo_col)}{excel_row_idx}"
                        ws.add_image(xl_img)
                        ws.row_dimensions[excel_row_idx].height = row_height
                        photo_cell_value = ""
                except Exception as e:
                    logger.warning("Фото в отчёт (result_id=%s): %s", row.get("result_id"), e)

            # Этаж, ряд, секция: из БД или парсинг из place_name (36.02.40.140.06.03 → этаж 2, ряд 40, секция 140)
            floor_val = row.get("floor")
            row_num_val = row.get("row_num")
            section_val = row.get("section")
            if (floor_val is None or row_num_val is None or section_val is None) and row.get("place_name"):
                pf, pr, ps = _parse_mx_code(row["place_name"])
                if floor_val is None:
                    floor_val = pf
                if row_num_val is None:
                    row_num_val = pr
                if section_val is None:
                    section_val = ps

            ws.append(
                [
                    created.isoformat(sep=" ") if created else "",
                    floor_val if floor_val is not None else "",
                    row_num_val if row_num_val is not None else "",
                    section_val if section_val is not None else "",
                    row["place_name"],
                    row["place_cod"],
                    _status_label(row.get("status")),
                    _error_description(row.get("discrepancy_reason"), row.get("comment")),
                    photo_cell_value,
                ]
            )

        ws.column_dimensions[get_column_letter(photo_col)].width = 18

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"history_not_ok_{badge}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        logger.exception("Ошибка экспорта истории сканов пользователя")
        return jsonify({'error': str(e)}), 500


@app.route('/api/user/session/start', methods=['POST'])
def start_user_session():
    """Начать новую сессию пользователя."""
    data = request.get_json()
    badge = data.get('badge')
    login_time = data.get('login_time')
    
    if not badge or not login_time:
        return jsonify({'error': 'Недостаточно данных'}), 400
    
    conn = None
    try:
        conn = get_db()
        with conn.cursor() as cur:
            # Закрываем все предыдущие активные сессии этого пользователя
            cur.execute("""
                UPDATE user_sessions 
                SET is_active = FALSE
                WHERE badge = %s AND is_active = TRUE
            """, (badge,))
            
            # Создаём новую сессию
            cur.execute("""
                INSERT INTO user_sessions (badge, login_time, is_active)
                VALUES (%s, %s, TRUE)
                RETURNING session_id
            """, (badge, login_time))
            
            session_id = cur.fetchone()['session_id']
            conn.commit()
            
        logger.info("Начата новая сессия для %s: session_id=%s", badge, session_id)
        
        return jsonify({
            'success': True,
            'session_id': session_id
        })
    
    except Exception as e:
        safe_rollback(conn)
        logger.exception("Ошибка создания сессии")
        return jsonify({'error': str(e)}), 500


@app.route('/api/user/session/update', methods=['POST'])
def update_user_session():
    """Обновить статистику текущей сессии."""
    data = request.get_json()
    badge = data.get('badge')
    stats = data.get('stats', {})
    
    if not badge:
        return jsonify({'error': 'Недостаточно данных'}), 400
    
    conn = None
    try:
        conn = get_db()
        with conn.cursor() as cur:
            # Обновляем активную сессию
            cur.execute("""
                UPDATE user_sessions 
                SET 
                    total_scanned = %s,
                    with_discrepancy = %s,
                    no_discrepancy = %s
                WHERE badge = %s AND is_active = TRUE
            """, (
                stats.get('total', 0),
                stats.get('withDiscrepancy', 0),
                stats.get('noDiscrepancy', 0),
                badge
            ))
            conn.commit()
        
        return jsonify({'success': True})
    
    except Exception as e:
        safe_rollback(conn)
        logger.exception("Ошибка обновления сессии")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/auth', methods=['POST'])
def admin_auth():
    """Авторизация администратора."""
    data = request.get_json()
    badge = data.get('badge', '').strip()
    password = data.get('password', '').strip()
    
    # Учетные данные администратора (в production лучше хранить в БД с хешированием)
    ADMIN_BADGE = 'ADMIN'
    ADMIN_PASSWORD = 'admin123'
    
    if not badge or not password:
        return jsonify({'success': False, 'error': 'Заполните все поля'}), 400
    
    if badge != ADMIN_BADGE or password != ADMIN_PASSWORD:
        return jsonify({'success': False, 'error': 'Неверный бэйдж или пароль'}), 401
    
    # Создаем ответ с установкой cookie
    response = jsonify({
        'success': True,
        'message': 'Авторизация успешна'
    })
    
    # Устанавливаем cookie на 8 часов
    response.set_cookie('admin_badge', badge, max_age=28800, httponly=True)
    response.set_cookie('employee_badge', '', max_age=0)
    
    return response


@app.route('/api/admin/logout', methods=['POST'])
def admin_logout():
    """Выход из админ-панели."""
    response = jsonify({'success': True})
    response.set_cookie('admin_badge', '', max_age=0)
    return response


@app.route('/api/place/<place_cod>', methods=['GET'])
def get_place(place_cod):
    """Получение данных о месте по place_cod (числовой ID или строковый mx_code)."""
    # Определяем тип идентификатора: числовой ID или строковый код МХ
    try:
        place_cod_int = int(place_cod)
        search_by_id = True
    except ValueError:
        # Если не число, значит это mx_code (например, "Э6.01.01.01")
        search_by_id = False
        place_cod_str = place_cod.strip().upper()
        
        # Проверяем корректность формата mx_code
        import re
        if not re.match(r'^[А-ЯЁA-Z0-9\.]+$', place_cod_str):
            return jsonify({'error': 'Некорректный формат кода МХ'}), 400

    try:
        conn = get_db()
        with conn.cursor() as cur:
            if search_by_id:
                # Поиск по числовому mx_id
                cur.execute(
                    """
                    SELECT 
                        mx_id as place_cod, 
                        mx_code as place_name, 
                        0 as qty_shk,
                        storage_type,
                        box_type,
                        dimensions,
                        category,
                        floor,
                        row_num,
                        section,
                        shelf,
                        cell,
                        current_volume,
                        current_occupancy,
                        updated_at
                    FROM warehouse_places
                    WHERE mx_id = %s
                    """,
                    (place_cod_int,),
                )
            else:
                # Поиск по строковому mx_code
                cur.execute(
                    """
                    SELECT 
                        mx_id as place_cod, 
                        mx_code as place_name, 
                        0 as qty_shk,
                        storage_type,
                        box_type,
                        dimensions,
                        category,
                        floor,
                        row_num,
                        section,
                        shelf,
                        cell,
                        current_volume,
                        current_occupancy,
                        updated_at
                    FROM warehouse_places
                    WHERE UPPER(mx_code) = %s
                    """,
                    (place_cod_str,),
                )
            
            row = cur.fetchone()

        if not row:
            return jsonify({'error': 'Место не найдено'}), 404

        # Определяем тип МХ — только Полка или Короб (без Микс)
        def _resolve_mx_type(storage_type, box_type, dimensions, category=None):
            for val in (storage_type, box_type, category):
                if not val:
                    continue
                s = str(val).lower()
                if 'короб' in s or 'box' in s:
                    return "Короб"
                if 'полка' in s or 'shelf' in s or 'стеллаж' in s:
                    return "Полка"
                # "Микс" не выводим — определяем по габаритам ниже
            if dimensions:
                try:
                    parts = str(dimensions).replace('х', 'x').replace('Х', 'x').split('x')
                    nums = [int(p.strip()) for p in parts if p.strip().isdigit()]
                    if nums:
                        max_d = max(nums)
                        if max_d > 900:
                            return "Полка"
                        return "Короб"  # до 900 мм — короб
                except (ValueError, TypeError):
                    pass
            # При неизвестном/микс без габаритов — по умолчанию короб
            if storage_type or box_type or category:
                return "Короб"
            return None

        mx_type = _resolve_mx_type(
            row.get('storage_type'), row.get('box_type'), row.get('dimensions'), row.get('category')
        ) or "—"
        
        return jsonify(
            {
                'place_cod': row['place_cod'],
                'place_name': row['place_name'],
                'qty_shk': row['qty_shk'],
                'mx_type': mx_type,
                'storage_type': row['storage_type'],
                'box_type': row['box_type'],
                'dimensions': row['dimensions'],
                'category': row['category'] or 'Не указана',
                'floor': row['floor'],
                'row_num': row['row_num'],
                'section': row['section'],
                'shelf': row['shelf'],
                'cell': row['cell'],
                'current_volume': float(row['current_volume']) if row['current_volume'] else None,
                'current_occupancy': row['current_occupancy'],
                'updated_at': row['updated_at'].isoformat() if row['updated_at'] else None,
            }
        )

    except Exception as e:
        logger.exception("Ошибка при получении данных о месте")
        return jsonify({'error': str(e)}), 500


@app.route('/api/scan/complete', methods=['POST'])
def complete_scan():
    """Фиксация результата сканирования сотрудником."""
    data = request.get_json() or {}
    badge = data.get('badge')
    place_cod = data.get('place_cod')
    fact_qty = data.get('fact_qty')
    status = data.get('status')
    comment = data.get('comment', '')
    discrepancy_reason = data.get('discrepancy_reason', '')
    # Поддержка нескольких фото: новое поле photos (список),
    # а поле photo оставляем для обратной совместимости (первое фото)
    photos = data.get('photos') or []
    photo_data = None
    photo_filename = None

    if not badge or not place_cod or not status:
        return jsonify({'error': 'Недостаточно данных'}), 400

    try:
        place_cod_int = int(place_cod)
    except (TypeError, ValueError):
        return jsonify({'error': 'Некорректный place_cod'}), 400

    # Декодируем одно или несколько фото
    decoded_photos = []
    import base64

    def _decode_photo(photo_str: str, index: int = 0):
        nonlocal photo_data, photo_filename
        if not photo_str:
            return
        try:
            if ',' in photo_str:
                header, encoded = photo_str.split(',', 1)
            else:
                header = ''
                encoded = photo_str
            raw = base64.b64decode(encoded)
            # Определяем расширение
            ext = "jpg"
            if 'png' in header:
                ext = "png"
            fname = f"{place_cod}_{index + 1}.{ext}"
            decoded_photos.append((raw, fname))
        except Exception as exc:
            logger.error("Ошибка декодирования фото #%s: %s", index + 1, exc)

    # Если пришёл массив photos — используем его, иначе старое поле photo
    if isinstance(photos, list) and photos:
        for idx, p in enumerate(photos):
            _decode_photo(p, idx)
    elif data.get('photo'):
        _decode_photo(data.get('photo'), 0)

    if decoded_photos:
        # Первое фото сохраняем также в inventory_results для обратной совместимости
        photo_data, photo_filename = decoded_photos[0]

    try:
        conn = get_db()
        place_row = None
        with conn.cursor() as cur:
            cur.execute("""
                SELECT mx_code as place_name, 0 as qty_shk
                FROM warehouse_places
                WHERE mx_id = %s
            """, (place_cod_int,))
            place_row = cur.fetchone()

            has_discrepancy = status != 'ok'
            qty_shk_db = place_row['qty_shk'] if place_row else None
            qty_fact_int = None
            if fact_qty is not None:
                try:
                    qty_fact_int = int(fact_qty)
                except (TypeError, ValueError):
                    qty_fact_int = None

            if qty_shk_db is not None and qty_fact_int is not None:
                try:
                    has_discrepancy = has_discrepancy or int(qty_shk_db) != qty_fact_int
                except ValueError:
                    pass

            cur.execute(
                """
                INSERT INTO inventory_results
                (badge, place_cod, place_name, qty_shk_db, qty_shk_fact, status, has_discrepancy, 
                 photo_data, photo_filename, discrepancy_reason, comment)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING result_id, created_at
                """,
                (
                    badge,
                    place_cod_int,
                    place_row['place_name'] if place_row else None,
                    qty_shk_db,
                    qty_fact_int,
                    status,
                    has_discrepancy,
                    psycopg2.Binary(photo_data) if photo_data else None,
                    photo_filename,
                    discrepancy_reason if discrepancy_reason else None,
                    comment if comment else None,
                ),
            )
            inserted = cur.fetchone()

            # Сохраняем все фото в отдельной таблице, если они есть
            if decoded_photos:
                for raw, fname in decoded_photos:
                    cur.execute(
                        """
                        INSERT INTO inventory_result_photos
                        (result_id, badge, place_cod, photo_data, photo_filename)
                        VALUES (%s, %s, %s, %s, %s)
                        """,
                        (
                            inserted["result_id"],
                            badge,
                            place_cod_int,
                            psycopg2.Binary(raw),
                            fname,
                        ),
                    )

            conn.commit()

        logger.info("Скан сохранен: badge=%s place=%s status=%s", badge, place_cod, status)
        return jsonify({
            'success': True,
            'result': {
                'id': inserted['result_id'],
                'place_cod': place_cod_int,
                'place_name': place_row['place_name'] if place_row else None,
                'qty_db': qty_shk_db,
                'qty_fact': qty_fact_int,
                'status': status,
                'has_discrepancy': has_discrepancy,
                'has_photo': photo_data is not None,
                'comment': comment,
                'created_at': inserted['created_at'].isoformat() if inserted and inserted['created_at'] else None
            }
        })
    except psycopg2.Error as e:
        if conn and not conn.closed:
            try:
                conn.rollback()
            except:
                pass
        logger.exception("Ошибка БД при сохранении скана")
        return jsonify({'error': f'Ошибка базы данных: {str(e)}'}), 500
    except Exception as e:
        if conn and not conn.closed:
            try:
                conn.rollback()
            except:
                pass
        logger.exception("Ошибка при сохранении скана")
        return jsonify({'error': str(e)}), 500


@app.route('/api/export', methods=['POST'])
def export_results():
    """Экспорт результатов инвентаризации в Excel и сохранение в БД."""
    data = request.get_json()
    badge = data.get('badge', 'unknown')
    results = data.get('results', [])
    
    if not results:
        return jsonify({'error': 'Нет данных для экспорта'}), 400
    
    conn = None
    try:
        # Сохраняем результаты в БД
        conn = get_db()
        with conn.cursor() as cur:
            for item in results:
                # Декодируем фото если есть
                photo_data = None
                photo_filename = None
                if item.get('photo'):
                    import base64
                    try:
                        # Формат: "data:image/jpeg;base64,..."
                        photo_str = item.get('photo')
                        if ',' in photo_str:
                            header, encoded = photo_str.split(',', 1)
                            photo_data = base64.b64decode(encoded)
                            # Определяем расширение из заголовка
                            if 'jpeg' in header or 'jpg' in header:
                                photo_filename = f"{item.get('place_cod')}.jpg"
                            elif 'png' in header:
                                photo_filename = f"{item.get('place_cod')}.png"
                    except Exception as e:
                        logger.error("Ошибка декодирования фото: %s", e)
                
                cur.execute("""
                    INSERT INTO inventory_results 
                    (badge, place_cod, place_name, qty_shk_db, qty_shk_fact, status, has_discrepancy, photo_data, photo_filename)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    badge,
                    item.get('place_cod'),
                    item.get('place_name'),
                    item.get('qty_shk_db'),
                    item.get('qty_shk_fact'),
                    item.get('status'),
                    item.get('has_discrepancy', False),
                    photo_data,
                    photo_filename
                ))
            conn.commit()
        
        logger.info("Сохранено %d записей в БД для сотрудника %s", len(results), badge)
        
        # Подтягиваем тип/этаж/ряд/секция из warehouse_places для отчёта
        place_cods = [item.get('place_cod') for item in results if item.get('place_cod') is not None]
        place_info = {}
        if place_cods:
            with conn.cursor() as cur2:
                cur2.execute("""
                    SELECT mx_id, storage_type, box_type, category, dimensions, floor, row_num, section
                    FROM warehouse_places
                    WHERE mx_id = ANY(%s)
                """, (place_cods,))
                for r in cur2.fetchall():
                    place_info[r['mx_id']] = r

        def _fmt_type(st, box_type=None, category=None, dimensions=None):
            """Только Полка или Короб."""
            for val in (st, box_type, category):
                if not val:
                    continue
                s = str(val).lower()
                if "короб" in s or "box" in s:
                    return "Короб"
                if "полка" in s or "shelf" in s or "стеллаж" in s:
                    return "Полка"
            if dimensions:
                try:
                    parts = str(dimensions).replace("х", "x").replace("Х", "x").split("x")
                    nums = [int(p.strip()) for p in parts if p.strip().isdigit()]
                    if nums:
                        if max(nums) > 900:
                            return "Полка"
                        return "Короб"
                except (ValueError, TypeError):
                    pass
            if st or box_type or category:
                return "Короб"
            return ""

        # Создаём Excel файл
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Инвентаризация"
        
        # Заголовки: этаж, ряд, секция, код МХ, ID места, статус, время
        headers = [
            'Этаж', 'Ряд', 'Секция',
            'Код МХ', 'ID места', 'Статус', 'Время'
        ]
        ws.append(headers)
        
        # Жирный шрифт для заголовков
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')
            cell.fill = header_fill
        
        # Цвета для подсветки
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Красный фон
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Зеленый фон
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Желтый фон
        num_cols = len(headers)
        
        # Данные
        row_num = 2
        for item in results:
            qty_db = item.get('qty_shk_db')
            qty_fact = item.get('qty_shk_fact')
            status = item.get('status', '')
            has_discrepancy = item.get('has_discrepancy', False)
            pc = item.get('place_cod')
            wp = place_info.get(pc) or {}
            
            ws.append([
                wp.get('floor') if wp.get('floor') is not None else "",
                wp.get('row_num') if wp.get('row_num') is not None else "",
                wp.get('section') or "",
                item.get('place_name'),
                pc,
                status,
                item.get('timestamp')
            ])
            
            # Подсветка строк с расхождениями
            if has_discrepancy and status != 'OK':
                for col in range(1, num_cols + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.fill = red_fill
            elif status == 'OK':
                for col in range(1, num_cols + 1):
                    ws.cell(row=row_num, column=col).fill = green_fill
            elif has_discrepancy:
                for col in range(1, num_cols + 1):
                    ws.cell(row=row_num, column=col).fill = yellow_fill
            
            row_num += 1
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Сохранение в память (BytesIO)
        from io import BytesIO
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"inventory_{badge}_{timestamp}.xlsx"
        
        file_stream = BytesIO()
        wb.save(file_stream)
        file_data = file_stream.getvalue()
        
        logger.info("Экспорт создан: %s (%d записей)", filename, len(results))
        
        # Подсчет статистики
        total_scanned = len(results)
        with_discrepancy = sum(1 for item in results if item.get('has_discrepancy', False))
        no_discrepancy = total_scanned - with_discrepancy
        
        # Сохраняем отчет в БД для админ-панели
        report_id = save_report(badge, file_data, filename, total_scanned, with_discrepancy, no_discrepancy)
        
        if report_id:
            return jsonify({
                'success': True,
                'message': 'Отчет отправлен в админ-панель',
                'filename': filename,
                'report_id': report_id
            })
        else:
            return jsonify({'error': 'Ошибка сохранения отчета'}), 500
    
    except psycopg2.Error as e:
        safe_rollback(conn)
        logger.exception("Ошибка БД при экспорте в Excel")
        return jsonify({'error': f'Ошибка базы данных: {str(e)}'}), 500
    except Exception as e:
        safe_rollback(conn)
        logger.exception("Ошибка при экспорте в Excel")
        return jsonify({'error': str(e)}), 500


@app.route('/api/task/new', methods=['POST'])
def get_new_task():
    """Получить новое задание для сотрудника."""
    try:
        data = request.get_json()
        badge = data.get('badge', '')
        zone_size = data.get('zone_size', 50)
        
        logger.info("Запрос нового задания от %s, размер: %d", badge, zone_size)
        
        # Очищаем просроченные задания
        cleanup_expired_tasks()
        
        # Получаем список занятых зон
        occupied_zones = get_occupied_zones()
        
        conn = get_db()
        max_attempts = 10  # Максимум попыток найти свободную зону
        
        for attempt in range(max_attempts):
            with conn.cursor() as cur:
                # Ищем случайную зону, которая не занята
                cur.execute("""
                    WITH random_place AS (
                        SELECT mx_code
                        FROM warehouse_places
                        WHERE mx_code IS NOT NULL AND mx_code != ''
                        ORDER BY RANDOM()
                        LIMIT 1
                    )
                    SELECT 
                        w.mx_id as place_cod,
                        w.mx_code as place_name,
                        0 as qty_shk
                    FROM warehouse_places w, random_place r
                    WHERE w.mx_code LIKE SUBSTRING(r.mx_code, 1, 9) || '%%'
                        AND w.mx_code IS NOT NULL
                    ORDER BY w.mx_code
                    LIMIT %s
                """, (zone_size,))
                
                rows = cur.fetchall()
            
            if not rows:
                continue
            
            places = [dict(row) for row in rows]
            zone_prefix = places[0]['place_name'][:9] if places[0]['place_name'] else None
            
            if not zone_prefix:
                continue
            
            # Проверяем, не занята ли эта зона
            if zone_prefix in occupied_zones:
                logger.info("Попытка %d: зона %s занята, ищем другую", attempt + 1, zone_prefix)
                continue
            
            # Пытаемся зарезервировать зону
            if reserve_zone(zone_prefix, badge, hours=2):
                task = {
                    'zone': zone_prefix,
                    'total_places': len(places),
                    'places': places,
                    'reserved': True
                }
                
                logger.info(
                    "✅ Задание создано: зона=%s, мест=%d, сотрудник=%s",
                    task['zone'],
                    task['total_places'],
                    badge
                )
                
                return jsonify({
                    'success': True,
                    'task': task,
                    'timestamp': datetime.now().isoformat(),
                    'expires_in_hours': 2
                })
            else:
                # Зона уже занята другим сотрудником
                occupied_zones.add(zone_prefix)
                continue
        
        # Если не нашли свободную зону за max_attempts попыток
        return jsonify({
            'success': False,
            'error': 'Все доступные зоны заняты. Попробуйте позже.'
        }), 503
    
    except Exception as e:
        logger.exception("Ошибка при создании задания")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/task/complete', methods=['POST'])
def complete_task():
    """Завершить задание и освободить зону."""
    conn = None
    try:
        data = request.get_json()
        badge = data.get('badge', '')
        zone = data.get('zone', '')
        
        if not badge or not zone:
            return jsonify({'error': 'Не указан badge или zone'}), 400
        
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE active_tasks 
                SET status = 'completed'
                WHERE zone_prefix = %s 
                    AND badge = %s
                    AND status = 'active'
            """, (zone, badge))
            
            updated = cur.rowcount
            conn.commit()
        
        if updated > 0:
            logger.info("✅ Задание завершено: зона=%s, сотрудник=%s", zone, badge)
            return jsonify({
                'success': True,
                'message': 'Задание завершено'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Задание не найдено или уже завершено'
            })
    
    except Exception as e:
        safe_rollback(conn)
        logger.exception("Ошибка при завершении задания")
        return jsonify({'error': str(e)}), 500


@app.route('/api/tasks/active', methods=['GET'])
def get_active_tasks():
    """Получить список активных заданий (для администратора)."""
    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT 
                    task_id,
                    zone_prefix,
                    badge,
                    assigned_at,
                    expires_at,
                    status,
                    EXTRACT(EPOCH FROM (expires_at - CURRENT_TIMESTAMP))/3600 as hours_left
                FROM active_tasks
                WHERE status = 'active'
                    AND expires_at > CURRENT_TIMESTAMP
                ORDER BY assigned_at DESC
            """)
            
            tasks = []
            for row in cur.fetchall():
                tasks.append({
                    'task_id': row['task_id'],
                    'zone': row['zone_prefix'],
                    'badge': row['badge'],
                    'assigned_at': row['assigned_at'].isoformat(),
                    'expires_at': row['expires_at'].isoformat(),
                    'hours_left': round(row['hours_left'], 1)
                })
        
        return jsonify({
            'success': True,
            'count': len(tasks),
            'tasks': tasks
        })
    
    except Exception as e:
        logger.exception("Ошибка при получении активных заданий")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/tasks/assign', methods=['POST'])
def admin_assign_task():
    """Назначить зону вручную (администратор)."""
    admin_badge = request.cookies.get('admin_badge')
    if not admin_badge or admin_badge != 'ADMIN':
        return jsonify({'error': 'Доступ запрещен'}), 403

    data = request.get_json() or {}
    badge = data.get('badge')
    zone_prefix = data.get('zone')
    hours = data.get('hours', 2)

    if not badge or not zone_prefix:
        return jsonify({'error': 'Укажите badge и зону'}), 400

    conn = None
    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO active_tasks (zone_prefix, badge, expires_at)
                VALUES (%s, %s, CURRENT_TIMESTAMP + INTERVAL '%s hours')
                RETURNING task_id, assigned_at, expires_at
            """, (zone_prefix, badge, hours))
            task = cur.fetchone()
            conn.commit()

        return jsonify({
            'success': True,
            'task': {
                'task_id': task['task_id'],
                'zone': zone_prefix,
                'badge': badge,
                'assigned_at': task['assigned_at'].isoformat() if task['assigned_at'] else None,
                'expires_at': task['expires_at'].isoformat() if task['expires_at'] else None
            }
        })
    except Exception as e:
        safe_rollback(conn)
        logger.exception("Ошибка назначения зоны")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/tasks/extend', methods=['POST'])
def admin_extend_task():
    """Продлить время активного задания."""
    admin_badge = request.cookies.get('admin_badge')
    if not admin_badge or admin_badge != 'ADMIN':
        return jsonify({'error': 'Доступ запрещен'}), 403

    data = request.get_json() or {}
    task_id = data.get('task_id')
    hours = data.get('hours', 1)

    if not task_id:
        return jsonify({'error': 'Не передан task_id'}), 400

    conn = None
    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE active_tasks
                SET expires_at = expires_at + INTERVAL '%s hours'
                WHERE task_id = %s AND status = 'active'
                RETURNING task_id, zone_prefix, badge, expires_at
            """, (hours, task_id))
            updated = cur.fetchone()
            conn.commit()

        if not updated:
            return jsonify({'error': 'Задача не найдена или уже закрыта'}), 404

        return jsonify({
            'success': True,
            'task': {
                'task_id': updated['task_id'],
                'zone': updated['zone_prefix'],
                'badge': updated['badge'],
                'expires_at': updated['expires_at'].isoformat() if updated['expires_at'] else None
            }
        })
    except Exception as e:
        safe_rollback(conn)
        logger.exception("Ошибка продления задания")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/tasks/close', methods=['POST'])
def admin_close_task():
    """Закрыть активное задание."""
    admin_badge = request.cookies.get('admin_badge')
    if not admin_badge or admin_badge != 'ADMIN':
        return jsonify({'error': 'Доступ запрещен'}), 403

    data = request.get_json() or {}
    task_id = data.get('task_id')

    if not task_id:
        return jsonify({'error': 'Не передан task_id'}), 400

    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE active_tasks
                SET status = 'completed'
                WHERE task_id = %s AND status = 'active'
                RETURNING zone_prefix, badge
            """, (task_id,))
            updated = cur.fetchone()
            conn.commit()

        if not updated:
            return jsonify({'error': 'Задача уже закрыта или не найдена'}), 404

        return jsonify({
            'success': True,
            'message': f"Задача {task_id} закрыта"
        })
    except Exception as e:
        logger.exception("Ошибка закрытия задания")
        return jsonify({'error': str(e)}), 500


@app.route('/api/tasks/suggestions', methods=['GET'])
def get_task_suggestions():
    """Ближайшие МХ: отталкиваясь от последнего отсканированного места сотрудника."""
    badge = request.args.get('badge')
    near = request.args.get('near')  # опционально: mx_code референсного МХ
    conn = None
    try:
        conn = get_db()
        ref_floor = ref_row = ref_section = None

        with conn.cursor() as cur:
            # Референс: последнее отсканированное место сотрудника или переданный near
            if near:
                cur.execute("""
                    SELECT floor, row_num, section FROM warehouse_places
                    WHERE UPPER(TRIM(mx_code)) = UPPER(TRIM(%s))
                    LIMIT 1
                """, (near,))
                row = cur.fetchone()
                if row:
                    ref_floor = row.get('floor')
                    ref_row = row.get('row_num')
                    ref_section = row.get('section')
            elif badge:
                cur.execute("""
                    SELECT ir.place_cod, ir.place_name
                    FROM inventory_results ir
                    WHERE ir.badge = %s AND (ir.place_cod IS NOT NULL OR ir.place_name IS NOT NULL)
                    ORDER BY ir.created_at DESC NULLS LAST
                    LIMIT 1
                """, (badge,))
                last_scan = cur.fetchone()
                if last_scan:
                    pc = last_scan.get('place_cod')
                    pn = last_scan.get('place_name')
                    if pc is not None:
                        cur.execute(
                            "SELECT floor, row_num, section FROM warehouse_places WHERE mx_id = %s LIMIT 1",
                            (pc,)
                        )
                    elif pn:
                        cur.execute(
                            "SELECT floor, row_num, section FROM warehouse_places WHERE UPPER(TRIM(mx_code)) = UPPER(TRIM(%s)) LIMIT 1",
                            (pn,)
                        )
                    else:
                        cur.execute("SELECT 1 WHERE FALSE")
                    ref_row_db = cur.fetchone()
                    if ref_row_db:
                        ref_floor = ref_row_db.get('floor')
                        ref_row = ref_row_db.get('row_num')
                        ref_section = ref_row_db.get('section')

            # МХ с приоритетом (есть расхождения в инвентаризации)
            cur.execute("""
                SELECT ir.place_name AS mx_code
                FROM inventory_results ir
                WHERE ir.has_discrepancy AND ir.place_name IS NOT NULL AND ir.place_name != ''
                GROUP BY ir.place_name
                ORDER BY MAX(ir.created_at) DESC NULLS LAST
                LIMIT 5
            """)
            priority_mx = {row['mx_code'].strip().upper(): True for row in cur.fetchall() if row.get('mx_code')}

            # Ближайшие МХ: сортировка по расстоянию от референса
            rf = ref_floor if ref_floor is not None else 0
            rr = ref_row if ref_row is not None else 0
            rs = ref_section if ref_section is not None else 0

            cur.execute("""
                WITH places AS (
                    SELECT mx_code, floor, row_num, section,
                           COALESCE(floor, 0) AS f,
                           COALESCE(row_num, 0) AS r,
                           COALESCE(section, 0) AS s
                    FROM warehouse_places
                    WHERE mx_code IS NOT NULL AND mx_code != ''
                )
                SELECT mx_code,
                       ABS(f - %s) + ABS(r - %s) + ABS(s - %s) AS dist
                FROM places
                ORDER BY dist ASC, mx_code ASC
                LIMIT 20
            """, (rf, rr, rs))

            nearest = [dict(row) for row in cur.fetchall()]

        suggestions = []
        seen_mx = set()
        for row in nearest:
            mc = (row.get('mx_code') or '').strip()
            if not mc or mc.upper() in seen_mx:
                continue
            seen_mx.add(mc.upper())
            suggestions.append({
                'mx_code': mc,
                'zone': mc[:9] if len(mc) >= 9 else mc,  # для обратной совместимости
                'highlight': mc.upper() in priority_mx,
            })
            if len(suggestions) >= 12:
                break

        return jsonify({
            'success': True,
            'suggestions': suggestions,
            'badge': badge,
        })
    except Exception as e:
        logger.exception("Ошибка генерации рекомендаций")
        safe_rollback(conn)
        return jsonify({'error': str(e)}), 500


@app.route('/api/sync', methods=['GET'])
def sync_data():
    """Синхронизация всех данных для offline работы."""
    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT mx_id as place_cod, mx_code as place_name, 0 as qty_shk
                FROM warehouse_places
                ORDER BY mx_id
            """)
            rows = cur.fetchall()
        
        data = [dict(row) for row in rows]
        
        logger.info("Синхронизация: отправлено %d записей", len(data))
        
        return jsonify({
            'success': True,
            'count': len(data),
            'data': data,
            'timestamp': datetime.now().isoformat()
        })
    
    except Exception as e:
        logger.exception("Ошибка при синхронизации данных")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/analytics', methods=['GET'])
def get_admin_analytics():
    """Получить данные для графиков."""
    # Проверка авторизации
    admin_badge = request.cookies.get('admin_badge')
    if not admin_badge or admin_badge != 'ADMIN':
        return jsonify({'error': 'Доступ запрещен'}), 403
    
    try:
        conn = get_db()
        with conn.cursor() as cur:
            # Динамика по дням (последние 7 дней)
            cur.execute("""
                SELECT 
                    DATE(created_at) as date,
                    COUNT(*) as total,
                    SUM(CASE WHEN has_discrepancy THEN 1 ELSE 0 END) as errors,
                    SUM(CASE WHEN NOT has_discrepancy THEN 1 ELSE 0 END) as ok
                FROM inventory_results
                WHERE created_at >= CURRENT_DATE - INTERVAL '7 days'
                GROUP BY DATE(created_at)
                ORDER BY date
            """)
            daily_stats = []
            for row in cur.fetchall():
                daily_stats.append({
                    'date': row['date'].isoformat(),
                    'total': row['total'],
                    'errors': row['errors'],
                    'ok': row['ok']
                })
            
            # Распределение сотрудников по точности
            cur.execute("""
                SELECT 
                    badge,
                    COUNT(*) as total,
                    SUM(CASE WHEN NOT has_discrepancy THEN 1 ELSE 0 END) as ok
                FROM inventory_results
                GROUP BY badge
                HAVING COUNT(*) >= 10
                ORDER BY badge
                LIMIT 200
            """)
            accuracy_stats = []
            for row in cur.fetchall():
                accuracy = round((row['ok'] / row['total']) * 100, 1) if row['total'] > 0 else 0
                accuracy_stats.append({
                    'badge': row['badge'],
                    'accuracy': accuracy
                })
            
            # ТОП проблемных мест
            cur.execute("""
                SELECT 
                    place_cod,
                    place_name,
                    COUNT(*) as scan_count,
                    SUM(CASE WHEN has_discrepancy THEN 1 ELSE 0 END) as error_count
                FROM inventory_results
                GROUP BY place_cod, place_name
                HAVING SUM(CASE WHEN has_discrepancy THEN 1 ELSE 0 END) > 0
                ORDER BY error_count DESC, scan_count DESC
                LIMIT 20
            """)
            problem_zones = []
            for row in cur.fetchall():
                problem_zones.append({
                    'place_cod': row['place_cod'],
                    'place_name': row['place_name'],
                    'scan_count': row['scan_count'],
                    'error_count': row['error_count'],
                    'error_rate': round((row['error_count'] / row['scan_count']) * 100, 1)
                })
        
        return jsonify({
            'success': True,
            'daily_stats': daily_stats,
            'accuracy_stats': accuracy_stats,
            'problem_zones': problem_zones
        })
    
    except Exception as e:
        logger.exception("Ошибка при получении аналитики")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/stats', methods=['GET'])
def get_admin_stats():
    """Получить общую статистику для админ-панели."""
    # Проверка авторизации
    admin_badge = request.cookies.get('admin_badge')
    if not admin_badge or admin_badge != 'ADMIN':
        return jsonify({'error': 'Доступ запрещен'}), 403
    
    try:
        conn = get_db()
        with conn.cursor() as cur:
            # Общая статистика
            cur.execute("""
                SELECT 
                    COUNT(*) as total_scanned,
                    COUNT(DISTINCT badge) as total_employees,
                    SUM(CASE WHEN has_discrepancy THEN 1 ELSE 0 END) as with_discrepancy,
                    SUM(CASE WHEN NOT has_discrepancy THEN 1 ELSE 0 END) as no_discrepancy,
                    COUNT(DISTINCT place_cod) as unique_places
                FROM inventory_results
            """)
            overall = dict(cur.fetchone())
            
            # Статистика по сотрудникам (ТОП 20 по количеству сканов)
            cur.execute("""
                SELECT 
                    badge,
                    COUNT(*) as scanned,
                    SUM(CASE WHEN has_discrepancy THEN 1 ELSE 0 END) as discrepancies,
                    MIN(created_at) as first_scan,
                    MAX(created_at) as last_scan
                FROM inventory_results
                GROUP BY badge
                ORDER BY scanned DESC
                LIMIT 20
            """)
            
            employee_rows = cur.fetchall()
            badges = [row['badge'] for row in employee_rows]

            # Общее время работы по всем этим бэйджам одной выборкой
            durations = {}
            if badges:
                cur.execute(
                    """
                    SELECT badge, COALESCE(SUM(session_duration), 0) AS total_seconds
                    FROM user_sessions
                    WHERE badge = ANY(%s) AND session_duration IS NOT NULL
                    GROUP BY badge
                    """,
                    (badges,),
                )
                for d_row in cur.fetchall():
                    durations[d_row['badge']] = d_row['total_seconds'] or 0

            employees = []
            for row in employee_rows:
                scanned = row['scanned']
                badge = row['badge']
                total_seconds = durations.get(badge, 0)
                total_hours = total_seconds / 3600 if total_seconds else 0
                speed = round(scanned / total_hours, 1) if total_hours > 0 else 0
                
                employees.append({
                    'badge': badge,
                    'scanned': scanned,
                    'discrepancies': row['discrepancies'],
                    'accuracy': round((1 - row['discrepancies'] / scanned) * 100, 1) if scanned > 0 else 100,
                    'total_hours': round(total_hours, 1),
                    'speed': speed,  # сканов в час
                    'first_scan': row['first_scan'].isoformat() if row['first_scan'] else None,
                    'last_scan': row['last_scan'].isoformat() if row['last_scan'] else None
                })
            
            # Статистика по типам расхождений
            cur.execute("""
                SELECT 
                    status,
                    COUNT(*) as count
                FROM inventory_results
                WHERE has_discrepancy = TRUE
                GROUP BY status
                ORDER BY count DESC
            """)
            discrepancy_types = [dict(row) for row in cur.fetchall()]
            
            # Динамика по часам (последние 24 часа)
            cur.execute("""
                SELECT 
                    DATE_TRUNC('hour', created_at) as hour,
                    COUNT(*) as count
                FROM inventory_results
                WHERE created_at >= NOW() - INTERVAL '24 hours'
                GROUP BY DATE_TRUNC('hour', created_at)
                ORDER BY hour
            """)
            hourly_stats = []
            for row in cur.fetchall():
                hourly_stats.append({
                    'hour': row['hour'].isoformat() if row['hour'] else None,
                    'count': row['count']
                })
        
        return jsonify({
            'success': True,
            'overall': overall,
            'employees': employees,
            'discrepancy_types': discrepancy_types,
            'hourly_stats': hourly_stats
        })
    
    except Exception as e:
        logger.exception("Ошибка получения статистики")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/latest_scans', methods=['GET'])
def get_admin_latest_scans():
    """Последние результаты сканирования для галереи."""
    admin_badge = request.cookies.get('admin_badge')
    if not admin_badge or admin_badge != 'ADMIN':
        return jsonify({'error': 'Доступ запрещен'}), 403

    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT result_id, place_cod, place_name, badge, qty_shk_db, qty_shk_fact,
                       status, has_discrepancy, photo_filename,
                       created_at
                FROM inventory_results
                ORDER BY created_at DESC
                LIMIT 12
            """)
            rows = cur.fetchall()

        scans = []
        for row in rows:
            scans.append({
                'id': row['result_id'],
                'place_cod': row['place_cod'],
                'place_name': row['place_name'],
                'badge': row['badge'],
                'qty_db': row['qty_shk_db'],
                'qty_fact': row['qty_shk_fact'],
                'status': row['status'],
                'has_discrepancy': row['has_discrepancy'],
                'has_photo': bool(row['photo_filename']),
                'photo_url': f"/api/admin/photo/{row['result_id']}" if row['photo_filename'] else None,
                'created_at': row['created_at'].isoformat() if row['created_at'] else None
            })

        return jsonify({'success': True, 'scans': scans})
    except Exception as e:
        logger.exception("Ошибка получения последних сканов")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/place/<int:place_cod>/photos', methods=['GET'])
def get_place_photos(place_cod: int):
    """Получить все фото по указанному МХ (для админ-панели)."""
    admin_badge = request.cookies.get('admin_badge')
    if not admin_badge or admin_badge != 'ADMIN':
        return jsonify({'error': 'Доступ запрещен'}), 403

    try:
        conn = get_db()
        with conn.cursor() as cur:
            # Фото, хранящиеся прямо в inventory_results (старые записи и первое фото)
            cur.execute(
                """
                SELECT
                    result_id,
                    place_cod,
                    place_name,
                    badge,
                    qty_shk_db,
                    qty_shk_fact,
                    status,
                    has_discrepancy,
                    created_at,
                    photo_filename
                FROM inventory_results
                WHERE place_cod = %s
                  AND photo_filename IS NOT NULL
                """,
                (place_cod,),
            )
            result_photos = cur.fetchall()

            # Дополнительные фото из новой таблицы
            cur.execute(
                """
                SELECT
                    p.photo_id,
                    p.result_id,
                    p.place_cod,
                    r.place_name,
                    p.badge,
                    r.qty_shk_db,
                    r.qty_shk_fact,
                    r.status,
                    r.has_discrepancy,
                    p.created_at,
                    p.photo_filename
                FROM inventory_result_photos p
                LEFT JOIN inventory_results r ON r.result_id = p.result_id
                WHERE p.place_cod = %s
                """,
                (place_cod,),
            )
            extra_photos = cur.fetchall()

        photos = []
        # Старые / первые фото
        for row in result_photos:
            photos.append(
                {
                    "id": row["result_id"],
                    "place_cod": row["place_cod"],
                    "place_name": row["place_name"],
                    "badge": row["badge"],
                    "qty_db": row["qty_shk_db"],
                    "qty_fact": row["qty_shk_fact"],
                    "status": row["status"],
                    "has_discrepancy": row["has_discrepancy"],
                    "created_at": row["created_at"].isoformat() if row["created_at"] else None,
                    "photo_url": f"/api/admin/photo/{row['result_id']}",
                }
            )

        # Дополнительные фото
        for row in extra_photos:
            photos.append(
                {
                    "id": row["photo_id"],
                    "place_cod": row["place_cod"],
                    "place_name": row["place_name"],
                    "badge": row["badge"],
                    "qty_db": row["qty_shk_db"],
                    "qty_fact": row["qty_shk_fact"],
                    "status": row["status"],
                    "has_discrepancy": row["has_discrepancy"],
                    "created_at": row["created_at"].isoformat() if row["created_at"] else None,
                    "photo_url": f"/api/admin/photo_file/{row['photo_id']}",
                }
            )

        # Сортируем по дате (новые сверху)
        photos.sort(key=lambda x: x["created_at"] or "", reverse=True)

        return jsonify({"success": True, "photos": photos})
    except Exception as e:
        logger.exception("Ошибка получения фото по месту")
        return jsonify({"error": str(e)}), 500


@app.route('/api/admin/activity', methods=['GET'])
def get_admin_activity():
    """Получить ленту событий (сканы, отчеты)."""
    admin_badge = request.cookies.get('admin_badge')
    if not admin_badge or admin_badge != 'ADMIN':
        return jsonify({'error': 'Доступ запрещен'}), 403

    try:
        conn = get_db()
        events = []

        with conn.cursor() as cur:
            moscow_offset = timedelta(hours=3)
            cur.execute("""
                SELECT badge, place_cod, place_name, status, created_at
                FROM inventory_results
                ORDER BY created_at DESC
                LIMIT 30
            """)
            for row in cur.fetchall():
                created = row['created_at']
                if created:
                    created = created + moscow_offset
                place_display = row['place_name'] or f"ID:{row['place_cod']}"
                events.append({
                    'type': 'scan',
                    'message': f"{row['badge']} зафиксировал {place_display} ({row['status'] or 'без статуса'})",
                    'timestamp': created.isoformat() if created else None
                })

            cur.execute("""
                SELECT badge, filename, created_at
                FROM reports
                ORDER BY created_at DESC
                LIMIT 10
            """)
            for row in cur.fetchall():
                created = row['created_at']
                if created:
                    created = created + moscow_offset
                events.append({
                    'type': 'report',
                    'message': f"Отчет {row['filename']} отправлен ({row['badge']})",
                    'timestamp': created.isoformat() if created else None
                })

        events.sort(key=lambda item: item['timestamp'] or '', reverse=True)
        events = events[:30]

        return jsonify({'success': True, 'events': events})
    except Exception as e:
        logger.exception("Ошибка получения ленты событий")
        return jsonify({'error': str(e)}), 500


@app.route('/api/tickets', methods=['POST'])
def create_ticket():
    """Создать тикет/инцидент от сотрудника."""
    data = request.get_json() or {}
    badge = data.get('badge')
    description = data.get('description')
    place_cod = data.get('place_cod')
    priority = data.get('priority', 'medium')

    if not badge or not description:
        return jsonify({'error': 'Заполните описание и badge'}), 400

    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO tickets (badge, place_cod, description, priority)
                VALUES (%s, %s, %s, %s)
                RETURNING ticket_id, created_at
            """, (badge, place_cod, description, priority))
            ticket = cur.fetchone()
            conn.commit()

        return jsonify({
            'success': True,
            'ticket': {
                'id': ticket['ticket_id'],
                'created_at': ticket['created_at'].isoformat() if ticket['created_at'] else None
            }
        })
    except psycopg2.Error as e:
        if conn and not conn.closed:
            try:
                conn.rollback()
            except:
                pass
        logger.exception("Ошибка БД при создании тикета")
        return jsonify({'error': f'Ошибка базы данных: {str(e)}'}), 500
    except Exception as e:
        if conn and not conn.closed:
            try:
                conn.rollback()
            except:
                pass
        logger.exception("Ошибка создания тикета")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/tickets', methods=['GET'])
def get_tickets():
    admin_badge = request.cookies.get('admin_badge')
    if not admin_badge or admin_badge != 'ADMIN':
        return jsonify({'error': 'Доступ запрещен'}), 403

    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT t.ticket_id, t.badge, t.place_cod, t.description, t.priority, t.status,
                       t.created_at, t.resolved_at, t.resolver,
                       wp.mx_code as place_name
                FROM tickets t
                LEFT JOIN warehouse_places wp ON t.place_cod = wp.mx_id
                ORDER BY t.status DESC, t.created_at DESC
                LIMIT 100
            """)
            tickets = []
            for row in cur.fetchall():
                tickets.append({
                    'id': row['ticket_id'],
                    'badge': row['badge'],
                    'place_cod': row['place_cod'],
                    'place_name': row['place_name'],
                    'description': row['description'],
                    'priority': row['priority'],
                    'status': row['status'],
                    'created_at': row['created_at'].isoformat() if row['created_at'] else None,
                    'resolved_at': row['resolved_at'].isoformat() if row['resolved_at'] else None,
                    'resolver': row['resolver']
                })
        return jsonify({'success': True, 'tickets': tickets})
    except Exception as e:
        logger.exception("Ошибка получения тикетов")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/tickets/<int:ticket_id>/resolve', methods=['POST'])
def resolve_ticket(ticket_id):
    admin_badge = request.cookies.get('admin_badge')
    if not admin_badge or admin_badge != 'ADMIN':
        return jsonify({'error': 'Доступ запрещен'}), 403

    conn = None
    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE tickets
                SET status = 'resolved',
                    resolver = %s,
                    resolved_at = CURRENT_TIMESTAMP
                WHERE ticket_id = %s AND status != 'resolved'
                RETURNING ticket_id
            """, (admin_badge, ticket_id))
            updated = cur.fetchone()
            conn.commit()

        if not updated:
            return jsonify({'error': 'Тикет уже закрыт или не найден'}), 404

        return jsonify({'success': True})
    except Exception as e:
        safe_rollback(conn)
        logger.exception("Ошибка закрытия тикета")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/reports', methods=['GET'])
def get_reports():
    """Получить список всех отчетов."""
    # Проверка авторизации
    admin_badge = request.cookies.get('admin_badge')
    if not admin_badge or admin_badge != 'ADMIN':
        return jsonify({'error': 'Доступ запрещен'}), 403
    
    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT 
                    report_id,
                    badge,
                    filename,
                    total_scanned,
                    with_discrepancy,
                    no_discrepancy,
                    created_at,
                    downloaded_at,
                    downloaded_by
                FROM reports
                ORDER BY created_at DESC
                LIMIT 100
            """)
            
            reports = []
            for row in cur.fetchall():
                reports.append({
                    'report_id': row['report_id'],
                    'badge': row['badge'],
                    'filename': row['filename'],
                    'total_scanned': row['total_scanned'],
                    'with_discrepancy': row['with_discrepancy'],
                    'no_discrepancy': row['no_discrepancy'],
                    'created_at': row['created_at'].isoformat() if row['created_at'] else None,
                    'downloaded_at': row['downloaded_at'].isoformat() if row['downloaded_at'] else None,
                    'downloaded_by': row['downloaded_by']
                })
        
        return jsonify({
            'success': True,
            'reports': reports
        })
    
    except Exception as e:
        logger.exception("Ошибка получения отчетов")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/wh_ids', methods=['GET'])
def get_wh_ids():
    """Список wh_id (складов) для выгрузки отчёта по складу."""
    admin_badge = request.cookies.get('admin_badge')
    if not admin_badge or admin_badge != 'ADMIN':
        return jsonify({'error': 'Доступ запрещен'}), 403
    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT DISTINCT wh_id,
                       MAX(warehouse_name) FILTER (WHERE warehouse_name IS NOT NULL AND warehouse_name != '') AS warehouse_name
                FROM warehouse_places
                WHERE wh_id IS NOT NULL
                GROUP BY wh_id
                ORDER BY wh_id
            """)
            wh_ids = [
                {'wh_id': row['wh_id'], 'warehouse_name': row['warehouse_name'] or ''}
                for row in cur.fetchall()
            ]
        return jsonify({'success': True, 'wh_ids': wh_ids})
    except Exception as e:
        logger.exception("Ошибка получения списка wh_id")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/export/block', methods=['GET'])
def export_block_errors():
    """Выгрузка отчёта по складу (wh_id): все ошибки (расхождения) по местам с данным wh_id. Параметр: wh_id."""
    admin_badge = request.cookies.get('admin_badge')
    if not admin_badge or admin_badge != 'ADMIN':
        return jsonify({'error': 'Доступ запрещен'}), 403

    wh_id_str = request.args.get('wh_id') or request.args.get('block', '').strip()
    if not wh_id_str:
        return jsonify({'error': 'Укажите wh_id (например 335, 133)'}), 400
    try:
        wh_id = int(wh_id_str)
    except ValueError:
        return jsonify({'error': 'wh_id должен быть числом'}), 400

    try:
        conn = get_db()
        with conn.cursor() as cur:
            query = """
                SELECT 
                    ir.result_id,
                    ir.created_at,
                    ir.place_cod,
                    ir.place_name,
                    ir.badge,
                    ir.qty_shk_db,
                    ir.qty_shk_fact,
                    ir.status,
                    ir.has_discrepancy,
                    ir.photo_filename,
                    ir.discrepancy_reason,
                    ir.comment,
                    wp.storage_type,
                    wp.box_type,
                    wp.category,
                    wp.dimensions,
                    COALESCE(wp.floor, wp2.floor) AS floor,
                    COALESCE(wp.row_num, wp2.row_num) AS row_num,
                    COALESCE(wp.section, wp2.section) AS section
                FROM inventory_results ir
                LEFT JOIN warehouse_places wp ON wp.mx_id = ir.place_cod
                LEFT JOIN warehouse_places wp2 ON UPPER(TRIM(wp2.mx_code)) = UPPER(TRIM(ir.place_name))
                    AND ir.place_name IS NOT NULL AND ir.place_name != ''
                WHERE (ir.has_discrepancy = TRUE OR LOWER(COALESCE(ir.status, '')) <> 'ok')
                  AND (wp.wh_id = %s OR wp2.wh_id = %s)
            """
            cur.execute(query + " ORDER BY ir.place_name, ir.created_at DESC", (wh_id, wh_id))
            rows = cur.fetchall()

        if not rows:
            return jsonify({'error': f'По wh_id {wh_id} нет записей с расхождениями'}), 404

        def _status_label(s):
            if not s:
                return ""
            s = (s or "").lower()
            if s == "ok": return "Совпадает"
            if s == "error": return "Ошибка"
            if s == "shelf_error": return "Поломалось"
            if s == "missing": return "Отсутствует"
            return s

        def _err_desc(reason, comment):
            parts = []
            if reason:
                parts.append(str(reason).strip())
            if comment:
                parts.append(f"Коммент.: {str(comment).strip()}")
            return " | ".join(parts) if parts else "—"

        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.drawing.image import Image as XlImage
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Ошибки wh_id {wh_id}"

        headers = [
            "Дата/время", "Этаж", "Ряд", "Секция",
            "Код МХ", "ID места", "Сотрудник",
            "Статус", "Причина/коммент.", "Фото"
        ]
        ws.append(headers)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
            cell.fill = header_fill

        photo_col = len(headers)
        row_height = 75
        img_max_height = 60
        MAX_EMBEDDED_PHOTOS = 100

        for excel_row_idx, row in enumerate(rows, start=2):
            created = row["created_at"]
            has_photo = bool(row.get("photo_filename"))
            photo_data = None
            if has_photo and row.get("result_id") and (excel_row_idx - 2) < MAX_EMBEDDED_PHOTOS:
                with conn.cursor() as cur_ph:
                    cur_ph.execute(
                        "SELECT photo_data FROM inventory_results WHERE result_id = %s AND photo_data IS NOT NULL LIMIT 1",
                        (row["result_id"],),
                    )
                    ph_row = cur_ph.fetchone()
                    if ph_row:
                        photo_data = ph_row.get("photo_data")
                    if not photo_data:
                        cur_ph.execute(
                            "SELECT photo_data FROM inventory_result_photos WHERE result_id = %s AND photo_data IS NOT NULL LIMIT 1",
                            (row["result_id"],),
                        )
                        ph_row = cur_ph.fetchone()
                        if ph_row:
                            photo_data = ph_row.get("photo_data")

            photo_cell_value = "есть" if has_photo else ""
            if has_photo and photo_data and (excel_row_idx - 2) < MAX_EMBEDDED_PHOTOS:
                try:
                    raw = photo_data
                    if hasattr(raw, "tobytes"):
                        raw = raw.tobytes()
                    elif not isinstance(raw, bytes):
                        raw = bytes(raw)
                    if raw:
                        img_io = BytesIO(raw)
                        img_io.seek(0)
                        xl_img = XlImage(img_io)
                        if xl_img.height and xl_img.height > img_max_height:
                            ratio = img_max_height / xl_img.height
                            xl_img.height = img_max_height
                            xl_img.width = int(xl_img.width * ratio)
                        xl_img.anchor = f"{get_column_letter(photo_col)}{excel_row_idx}"
                        ws.add_image(xl_img)
                        ws.row_dimensions[excel_row_idx].height = row_height
                        photo_cell_value = ""
                except Exception as e:
                    logger.warning("Фото в отчёт по блоку (result_id=%s): %s", row.get("result_id"), e)

            floor_val = row.get("floor")
            row_num_val = row.get("row_num")
            section_val = row.get("section")
            if (floor_val is None or row_num_val is None or section_val is None) and row.get("place_name"):
                parts = str(row["place_name"]).strip().split(".")
                try:
                    if floor_val is None and len(parts) >= 2:
                        floor_val = int(parts[1])
                    if row_num_val is None and len(parts) >= 3:
                        row_num_val = int(parts[2])
                    if section_val is None and len(parts) >= 4:
                        section_val = int(parts[3])
                except (ValueError, IndexError):
                    pass

            ws.append([
                created.isoformat(sep=" ") if created else "",
                floor_val if floor_val is not None else "",
                row_num_val if row_num_val is not None else "",
                section_val if section_val is not None else "",
                row["place_name"],
                row["place_cod"],
                row.get("badge") or "",
                _status_label(row.get("status")),
                _err_desc(row.get("discrepancy_reason"), row.get("comment")),
                photo_cell_value,
            ])

        ws.column_dimensions[get_column_letter(photo_col)].width = 18

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"errors_wh_id_{wh_id}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        logger.exception("Ошибка выгрузки отчёта по блоку")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/reviews', methods=['GET'])
def get_quality_reviews():
    admin_badge = request.cookies.get('admin_badge')
    if not admin_badge or admin_badge != 'ADMIN':
        return jsonify({'error': 'Доступ запрещен'}), 403

    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT SUBSTRING(place_name, 1, 9) as zone_prefix,
                       COUNT(*) as scan_count,
                       SUM(CASE WHEN has_discrepancy THEN 1 ELSE 0 END) as errors,
                       MAX(created_at) as last_scan
                FROM inventory_results
                WHERE place_name IS NOT NULL
                GROUP BY zone_prefix
                ORDER BY errors DESC, scan_count DESC
                LIMIT 20
            """)
            aggregates = []
            for row in cur.fetchall():
                aggregates.append({
                    'zone': row['zone_prefix'],
                    'scan_count': row['scan_count'],
                    'errors': row['errors'],
                    'last_scan': row['last_scan'].isoformat() if row['last_scan'] else None
                })

            cur.execute("""
                SELECT review_id, zone_prefix, reviewer, status, summary, created_at
                FROM quality_reviews
                ORDER BY created_at DESC
                LIMIT 20
            """)
            reviews = []
            for row in cur.fetchall():
                reviews.append({
                    'id': row['review_id'],
                    'zone': row['zone_prefix'],
                    'reviewer': row['reviewer'],
                    'status': row['status'],
                    'summary': row['summary'],
                    'created_at': row['created_at'].isoformat() if row['created_at'] else None
                })

        return jsonify({'success': True, 'aggregates': aggregates, 'reviews': reviews})
    except Exception as e:
        logger.exception("Ошибка выдачи ревизий")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/reviews', methods=['POST'])
def create_quality_review():
    admin_badge = request.cookies.get('admin_badge')
    if not admin_badge or admin_badge != 'ADMIN':
        return jsonify({'error': 'Доступ запрещен'}), 403

    data = request.get_json() or {}
    zone = data.get('zone')
    status = data.get('status', 'planned')
    summary = data.get('summary')

    if not zone:
        return jsonify({'error': 'Укажите зону'}), 400

    conn = None
    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO quality_reviews (zone_prefix, reviewer, status, summary)
                VALUES (%s, %s, %s, %s)
                RETURNING review_id, created_at
            """, (zone, admin_badge, status, summary))
            review = cur.fetchone()
            conn.commit()

        return jsonify({
            'success': True,
            'review': {
                'id': review['review_id'],
                'created_at': review['created_at'].isoformat() if review['created_at'] else None
            }
        })
    except Exception as e:
        safe_rollback(conn)
        logger.exception("Ошибка создания ревизии")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/reports/<int:report_id>/download', methods=['GET'])
def download_report(report_id):
    """Скачать отчет."""
    # Проверка авторизации
    admin_badge = request.cookies.get('admin_badge')
    if not admin_badge or admin_badge != 'ADMIN':
        return jsonify({'error': 'Доступ запрещен'}), 403
    
    conn = None
    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT file_data, filename
                FROM reports
                WHERE report_id = %s
            """, (report_id,))
            
            row = cur.fetchone()
            if not row:
                return jsonify({'error': 'Отчет не найден'}), 404
            
            file_data = bytes(row['file_data'])
            filename = row['filename']
            
            # Отмечаем скачивание
            cur.execute("""
                UPDATE reports
                SET downloaded_at = NOW(), downloaded_by = %s
                WHERE report_id = %s AND downloaded_at IS NULL
            """, (admin_badge, report_id))
            conn.commit()
        
        # Отправляем файл
        from io import BytesIO
        return send_file(
            BytesIO(file_data),
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        safe_rollback(conn)
        logger.exception("Ошибка скачивания отчета")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/photo/<int:result_id>', methods=['GET'])
def get_result_photo(result_id):
    """Получить фото для результата инвентаризации."""
    # Проверка авторизации
    admin_badge = request.cookies.get('admin_badge')
    if not admin_badge or admin_badge != 'ADMIN':
        return jsonify({'error': 'Доступ запрещен'}), 403
    
    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT photo_data, photo_filename
                FROM inventory_results
                WHERE result_id = %s AND photo_data IS NOT NULL
                """,
                (result_id,),
            )
            row = cur.fetchone()

            if not row or not row["photo_data"]:
                return jsonify({"error": "Фото не найдено"}), 404

            # Определяем mimetype
            mimetype = "image/jpeg"
            if row["photo_filename"] and row["photo_filename"].endswith(".png"):
                mimetype = "image/png"

            from io import BytesIO

            return send_file(
                BytesIO(row["photo_data"]),
                mimetype=mimetype,
                as_attachment=False,
                download_name=row["photo_filename"] or "photo.jpg",
            )
    
    except Exception as e:
        logger.exception("Ошибка при получении фото")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/photo/<int:result_id>/download', methods=['GET'])
def download_result_photo(result_id: int):
    """Скачать фото как файл (для админа)."""
    admin_badge = request.cookies.get('admin_badge')
    if not admin_badge or admin_badge != 'ADMIN':
        return jsonify({'error': 'Доступ запрещен'}), 403

    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT photo_data, photo_filename
                FROM inventory_results
                WHERE result_id = %s AND photo_data IS NOT NULL
                """,
                (result_id,),
            )
            row = cur.fetchone()

        if not row or not row["photo_data"]:
            return jsonify({"error": "Фото не найдено"}), 404

        mimetype = "image/jpeg"
        filename = row["photo_filename"] or f"photo_{result_id}.jpg"
        if filename.endswith(".png"):
            mimetype = "image/png"

        from io import BytesIO

        return send_file(
            BytesIO(row["photo_data"]),
            mimetype=mimetype,
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        logger.exception("Ошибка при скачивании фото")
        return jsonify({"error": str(e)}), 500


@app.route('/api/admin/photo/<int:result_id>', methods=['DELETE'])
def delete_result_photo(result_id: int):
    """Удалить фото у результата инвентаризации (для админа)."""
    admin_badge = request.cookies.get('admin_badge')
    if not admin_badge or admin_badge != 'ADMIN':
        return jsonify({'error': 'Доступ запрещен'}), 403

    conn = None
    try:
        conn = get_db()
        with conn.cursor() as cur:
            # Обнуляем только фото, сам результат оставляем
            cur.execute(
                """
                UPDATE inventory_results
                SET photo_data = NULL,
                    photo_filename = NULL
                WHERE result_id = %s
                """,
                (result_id,),
            )

            if cur.rowcount == 0:
                return jsonify({'error': 'Результат не найден'}), 404

        conn.commit()
        logger.info("Фото для результата %s удалено администратором", result_id)
        return jsonify({'success': True})

    except Exception as e:
        safe_rollback(conn)
        logger.exception("Ошибка при удалении фото")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/employee/<badge>', methods=['GET'])
def get_employee_details(badge):
    """Получить детальную статистику по сотруднику."""
    # Проверка авторизации
    admin_badge = request.cookies.get('admin_badge')
    if not admin_badge or admin_badge != 'ADMIN':
        return jsonify({'error': 'Доступ запрещен'}), 403
    
    try:
        conn = get_db()
        with conn.cursor() as cur:
            # Последние сканирования
            cur.execute("""
                SELECT 
                    result_id,
                    place_cod,
                    place_name,
                    qty_shk_db,
                    qty_shk_fact,
                    status,
                    has_discrepancy,
                    photo_data,
                    created_at
                FROM inventory_results
                WHERE badge = %s
                ORDER BY created_at DESC
                LIMIT 100
            """, (badge,))
            
            scans = []
            for row in cur.fetchall():
                scans.append({
                    'id': row['result_id'],
                    'place_cod': row['place_cod'],
                    'place_name': row['place_name'],
                    'qty_db': row['qty_shk_db'],
                    'qty_fact': row['qty_shk_fact'],
                    'status': row['status'],
                    'has_discrepancy': row['has_discrepancy'],
                    'has_photo': row['photo_data'] is not None,
                    'created_at': row['created_at'].isoformat()
                })
        
        return jsonify({
            'success': True,
            'badge': badge,
            'scans': scans
        })
    
    except Exception as e:
        logger.exception("Ошибка получения данных сотрудника")
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    # Запуск сервера
    logger.info("Запуск веб-сервера на http://127.0.0.1:8001")
    app.run(host='0.0.0.0', port=8001, debug=True)

